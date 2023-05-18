from openpyxl import load_workbook
import pymysql

def carregar_dados_excel(arquivo):
    try:
        # Carrega o arquivo Excel
        workbook = load_workbook(arquivo)
        sheet = workbook.active

        # Conecta ao banco de dados
        conexao = pymysql.connect(
            host="localhost",
            user="root",
            password="",
            database="pedro"
        )

        cursor = conexao.cursor()

        # Percorre as linhas do arquivo Excel (ignorando a primeira linha de cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome_aluno = row[0]
            data = row[1]
            aula_efetuada = row[2]
            materia = row[3]
            informacoes_adicionais = row[4]

            # Insere os dados na tabela teste_aluno
            query = "INSERT INTO teste_aluno (nome_aluno, data, aula_efetuada, materia, informacoes_adicionais) VALUES (%s, %s, %s, %s, %s)"
            valores = (nome_aluno, data, aula_efetuada, materia, informacoes_adicionais)
            cursor.execute(query, valores)

        # Realiza o commit das alterações no banco de dados
        conexao.commit()

        # Fecha a conexão com o banco de dados
        cursor.close()
        conexao.close()

        return True

    except Exception as e:
        print("Erro ao carregar os dados:", str(e))
        return False


# Exemplo de uso
arquivo_excel = 'C:/Users/Pedro/Documents/dados.xlsx'
resultado = carregar_dados_excel(arquivo_excel)

if resultado:
    print("Dados carregados com sucesso!")
else:
    print("Falha ao carregar os dados.")
